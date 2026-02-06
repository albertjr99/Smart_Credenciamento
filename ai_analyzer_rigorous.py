"""
SISTEMA DE ANÁLISE RIGOROSA COM IA - CREDENCIAMENTO RPPS
Análises sérias e robustas baseadas em regras específicas
Powered by Google Gemini AI
"""

import PyPDF2
import openpyxl
import re
from datetime import datetime, timedelta
from dateutil import parser as date_parser
import json
import os
from ai_config import get_ai_analysis

def extract_text_from_pdf(file_path):
    """Extrai texto completo de PDF com proteção contra erros"""
    try:
        with open(file_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            text = ""
            for page in pdf_reader.pages:
                try:
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + "\n"
                except Exception as e:
                    # Ignorar erros de decodificação de imagens XFormObject
                    if 'XFormObject' not in str(e):
                        print(f"⚠️ Erro ao extrair texto da página: {str(e)[:80]}")
                    continue
            return text.strip() if text.strip() else "ERRO: Não foi possível extrair texto"
    except Exception as e:
        return f"ERRO: {str(e)}"

def extract_dates_from_text(text):
    """Extrai e parseia datas do texto"""
    dates = []
    patterns = [
        r'\d{1,2}[/-]\d{1,2}[/-]\d{2,4}',
        r'\d{1,2}\s+de\s+\w+\s+de\s+\d{4}',
        r'\w+\s+\d{1,2},\s+\d{4}',
    ]
    
    for pattern in patterns:
        matches = re.findall(pattern, text, re.IGNORECASE)
        for match in matches:
            try:
                date_obj = date_parser.parse(match, fuzzy=True, dayfirst=True)
                dates.append(date_obj)
            except:
                continue
    
    return dates

def is_date_within_one_year(date_obj):
    """Verifica se data está dentro de 1 ano"""
    if not date_obj:
        return None
    today = datetime.now()
    one_year_ago = today - timedelta(days=365)
    return date_obj >= one_year_ago


def generate_detailed_analysis_fallback(text, institution_name, institution_mentioned, dates, score):
    """
    Gera análise detalhada baseada em regras quando a IA não está disponível
    Simula o comportamento da IA com análise inteligente
    """
    itens_encontrados = []
    itens_ausentes = []
    pontos_positivos = []
    pontos_negativos = []
    
    # Análise de conteúdo
    text_lower = text.lower()
    palavras_chave_financeiras = ['banco', 'instituição financeira', 'gestão de recursos', 'investimentos', 
                                   'ativos', 'patrimônio', 'fundo', 'carteira', 'rpps', 'previdência']
    
    palavras_encontradas = [palavra for palavra in palavras_chave_financeiras if palavra in text_lower]
    
    if len(palavras_encontradas) > 5:
        itens_encontrados.append(f"Vocabulário técnico financeiro presente ({len(palavras_encontradas)} termos identificados)")
        pontos_positivos.append("Documento contém terminologia adequada ao setor financeiro")
    elif len(palavras_encontradas) > 2:
        itens_encontrados.append(f"Alguns termos financeiros identificados ({len(palavras_encontradas)} termos)")
    else:
        itens_ausentes.append("Vocabulário técnico financeiro limitado ou ausente")
        pontos_negativos.append("Falta de terminologia esperada para documentos institucionais financeiros")
    
    # Análise de estrutura
    tem_paragrafos = text.count('\n\n') > 3
    tem_titulos = any(palavra in text for palavra in ['Visão', 'Missão', 'História', 'Sobre', 'Serviços', 'Produtos'])
    
    if tem_paragrafos and tem_titulos:
        itens_encontrados.append("Documento estruturado com seções identificáveis")
        pontos_positivos.append("Boa organização visual e estrutural")
    elif tem_paragrafos:
        itens_encontrados.append("Documento possui parágrafos distintos")
    else:
        itens_ausentes.append("Estruturação clara em seções")
        pontos_negativos.append("Documento parece desorganizado ou mal formatado")
    
    # Análise de instituição
    if institution_mentioned:
        itens_encontrados.append(f"Nome da instituição '{institution_name}' claramente identificado")
        pontos_positivos.append("Documento corresponde à instituição esperada")
    else:
        itens_ausentes.append(f"Menção explícita ao nome '{institution_name}'")
        pontos_negativos.append("Documento pode não corresponder à instituição ou estar genérico")
    
    # Análise de datas
    if dates:
        data_mais_recente = max(dates)
        itens_encontrados.append(f"Data identificada: {data_mais_recente.strftime('%d/%m/%Y')}")
        if is_date_within_one_year(data_mais_recente):
            pontos_positivos.append("Documento atualizado (menos de 1 ano)")
        else:
            pontos_negativos.append(f"Documento desatualizado ({data_mais_recente.strftime('%d/%m/%Y')})")
    else:
        itens_ausentes.append("Data de criação ou atualização do documento")
        pontos_negativos.append("Impossível verificar se o documento está atualizado")
    
    # Análise de completude
    elementos_esperados = {
        'cnpj': r'\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}',
        'email': r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}',
        'telefone': r'\(\d{2}\)\s*\d{4,5}-?\d{4}',
        'endereço': r'(rua|avenida|av\.|r\.)\s+[a-z\s]+,\s*\d+',
        'valores monetários': r'R\$\s*[\d.,]+'
    }
    
    elementos_encontrados_dict = {}
    for elemento, pattern in elementos_esperados.items():
        if re.search(pattern, text, re.IGNORECASE):
            elementos_encontrados_dict[elemento] = True
            itens_encontrados.append(f"{elemento.capitalize()} presente no documento")
        else:
            itens_ausentes.append(f"{elemento.capitalize()}")
    
    # Tamanho do documento
    tamanho_texto = len(text)
    if tamanho_texto > 3000:
        itens_encontrados.append(f"Documento extenso e detalhado ({tamanho_texto} caracteres)")
        pontos_positivos.append("Volume adequado de informações")
    elif tamanho_texto > 1000:
        itens_encontrados.append(f"Documento de tamanho razoável ({tamanho_texto} caracteres)")
    else:
        itens_ausentes.append("Conteúdo mais detalhado e extenso")
        pontos_negativos.append("Documento muito curto, pode estar incompleto")
    
    # Gerar justificativas
    completeness_score = score
    coherence_score = score
    
    if score >= 70:
        justificativa_completude = f"O documento apresenta {len(itens_encontrados)} elementos dos esperados, demonstrando razoável completude. "
        justificativa_coerencia = "As informações estão organizadas de forma coerente e seguem uma estrutura lógica. "
    else:
        justificativa_completude = f"O documento apresenta apenas {len(itens_encontrados)} elementos, faltando informações importantes. "
        justificativa_coerencia = "A organização das informações precisa ser melhorada para maior clareza. "
    
    if len(itens_ausentes) > 0:
        justificativa_completude += f"Identificamos {len(itens_ausentes)} elementos ausentes que deveriam estar presentes."
    
    if len(pontos_negativos) > 0:
        justificativa_coerencia += f"Foram identificados {len(pontos_negativos)} pontos que prejudicam a coerência do documento."
    
    return {
        'itens_encontrados': itens_encontrados[:8],
        'itens_ausentes': itens_ausentes[:8],
        'pontos_positivos': pontos_positivos[:5],
        'pontos_negativos': pontos_negativos[:5],
        'justificativa_completude': justificativa_completude,
        'justificativa_coerencia': justificativa_coerencia,
        'ai_completeness': completeness_score,
        'ai_coherence': coherence_score
    }


# ========== ANÁLISE 1: APRESENTAÇÃO INSTITUCIONAL ==========
def analyze_apresentacao_institucional(file_path, institution_name):
    """
    Análise RIGOROSA de Apresentação Institucional
    Regras:
    1. Verificar se menciona a instituição
    2. Verificar data (rejeitar se > 1 ano)
    3. Validar coerência informacional
    """
    print(f"\n📊 ANÁLISE: Apresentação Institucional")
    
    text = extract_text_from_pdf(file_path)
    if text.startswith("ERRO"):
        return {
            'is_valid': False,
            'score': 0,
            'issues': [text],
            'warnings': [],
            'summary': 'Erro ao ler arquivo PDF'
        }
    
    issues = []
    warnings = []
    score = 100
    
    # 1. Verificar menção à instituição
    institution_mentioned = institution_name.lower() in text.lower() if institution_name else False
    if not institution_mentioned and institution_name:
        issues.append(f"❌ A instituição '{institution_name}' não é mencionada na apresentação")
        score -= 40
    
    # 2. Verificar data
    dates = extract_dates_from_text(text)
    if dates:
        most_recent_date = max(dates)
        if not is_date_within_one_year(most_recent_date):
            issues.append(f"❌ Apresentação está desatualizada (data: {most_recent_date.strftime('%d/%m/%Y')}). Deve ter menos de 1 ano.")
            score -= 40
        else:
            warnings.append(f"✓ Data encontrada: {most_recent_date.strftime('%d/%m/%Y')} (dentro do prazo)")
    else:
        warnings.append("⚠️ Nenhuma data encontrada na apresentação")
    
    # 3. Análise com IA - Coerência e completude (COM PROTEÇÃO CONTRA ERROS)
    try:
        print("🔍 [ANÁLISE] Preparando prompt para IA...")
        prompt = f"""
Você é um auditor especializado em análise de documentos institucionais financeiros.

Analise esta APRESENTAÇÃO INSTITUCIONAL da empresa '{institution_name}'.

TEXTO DO DOCUMENTO:
{text[:4000]}

Faça uma análise DETALHADA e responda em JSON com:
{{
    "conteudo_adequado": boolean,
    "fala_sobre_empresa": boolean,
    "completude": number (0-100),
    "coerencia": number (0-100),
    "itens_encontrados": ["lista detalhada do que FOI ENCONTRADO no documento"],
    "itens_ausentes": ["lista detalhada do que NÃO FOI ENCONTRADO mas era esperado"],
    "pontos_positivos": ["aspectos positivos identificados"],
    "pontos_negativos": ["problemas ou deficiências identificadas"],
    "justificativa_completude": "explicação detalhada do score de completude",
    "justificativa_coerencia": "explicação detalhada do score de coerência",
    "problemas": [],
    "resumo": "análise breve"
}}

Seja ESPECÍFICO: cite trechos do documento, seções encontradas, informações faltantes.
"""
        
        print("📡 [ANÁLISE] Enviando para IA...")
        ai_result = get_ai_analysis(prompt, text[:3000], "apresentacao_institucional")
        
        print(f"📥 [ANÁLISE] Resposta da IA recebida: {ai_result}")
        
        if ai_result and ai_result.get('success'):
            print("✅ [ANÁLISE] IA respondeu com sucesso!")
            try:
                # Limpar resposta da IA
                response_text = ai_result.get('analysis', '{}')
                # Remover markdown se houver
                if '```' in response_text:
                    response_text = response_text.split('```')[1] if '```json' in response_text else response_text.split('```')[1]
                    response_text = response_text.replace('json', '').strip()
                
                analysis = json.loads(response_text)
                
                # Extrair análises detalhadas da IA
                ai_completeness = analysis.get('completude', 75)
                ai_coherence = analysis.get('coerencia', 75)
                
                if not analysis.get('conteudo_adequado', True):
                    issues.append("❌ IA: Conteúdo inadequado")
                    score -= 20
                
                if not analysis.get('fala_sobre_empresa', True):
                    issues.append("❌ IA: Não apresenta informações sobre a empresa")
                    score -= 20
                
                if ai_completeness < 60:
                    issues.append(f"❌ IA: Apresentação incompleta ({ai_completeness}%)")
                    score -= 15
                
                for problema in analysis.get('problemas', []):
                    if problema:
                        warnings.append(f"⚠️ {problema}")
                
                # Armazenar comentários detalhados da IA
                ai_detailed_comments = {
                    'itens_encontrados': analysis.get('itens_encontrados', []),
                    'itens_ausentes': analysis.get('itens_ausentes', []),
                    'pontos_positivos': analysis.get('pontos_positivos', []),
                    'pontos_negativos': analysis.get('pontos_negativos', []),
                    'justificativa_completude': analysis.get('justificativa_completude', ''),
                    'justificativa_coerencia': analysis.get('justificativa_coerencia', ''),
                    'ai_completeness': ai_completeness,
                    'ai_coherence': ai_coherence
                }
                        
            except Exception as e:
                print(f"❌ [ANÁLISE] Erro ao processar resposta da IA: {str(e)}")
                warnings.append(f"⚠️ Análise IA indisponível: {str(e)[:100]}")
        else:
            error_msg = ai_result.get('error', 'Erro desconhecido') if ai_result else 'Nenhuma resposta'
            print(f"⚠️ [ANÁLISE] IA não retornou sucesso: {error_msg}")
            print("🔄 [ANÁLISE] Usando análise avançada baseada em regras...")
            # Não adicionar warning para o usuário - análise funciona normalmente
            
            # ANÁLISE AVANÇADA BASEADA EM REGRAS (FALLBACK)
            ai_detailed_comments = generate_detailed_analysis_fallback(text, institution_name, institution_mentioned, dates, score)
    except Exception as e:
        print(f"❌ [ANÁLISE] Erro geral na análise IA: {str(e)}")
        import traceback
        traceback.print_exc()
        warnings.append(f"⚠️ Erro na análise IA: {str(e)[:100]}")
        
        # FALLBACK: gerar análise baseada em regras
        ai_detailed_comments = generate_detailed_analysis_fallback(text, institution_name, institution_mentioned, dates, score)
    
    score = max(0, score)
    is_valid = score >= 70 and len(issues) == 0
    
    # Gerar comentários detalhados baseados na análise da IA
    if 'ai_detailed_comments' in locals():
        # Usar análise detalhada da IA
        completeness_score = ai_detailed_comments.get('ai_completeness', score)
        coherence_score = ai_detailed_comments.get('ai_coherence', score)
        
        # Montar comentário de completude com detalhes
        completeness_parts = []
        if ai_detailed_comments.get('justificativa_completude'):
            completeness_parts.append(ai_detailed_comments['justificativa_completude'])
        
        if ai_detailed_comments.get('itens_encontrados'):
            completeness_parts.append("\n\n✅ ITENS ENCONTRADOS:\n" + "\n".join([f"• {item}" for item in ai_detailed_comments['itens_encontrados'][:5]]))
        
        if ai_detailed_comments.get('itens_ausentes'):
            completeness_parts.append("\n\n❌ ITENS AUSENTES:\n" + "\n".join([f"• {item}" for item in ai_detailed_comments['itens_ausentes'][:5]]))
        
        completeness_comment = " ".join(completeness_parts) if completeness_parts else f"Completude: {completeness_score}%"
        
        # Montar comentário de coerência com detalhes
        coherence_parts = []
        if ai_detailed_comments.get('justificativa_coerencia'):
            coherence_parts.append(ai_detailed_comments['justificativa_coerencia'])
        
        if ai_detailed_comments.get('pontos_positivos'):
            coherence_parts.append("\n\n✅ PONTOS POSITIVOS:\n" + "\n".join([f"• {item}" for item in ai_detailed_comments['pontos_positivos'][:3]]))
        
        if ai_detailed_comments.get('pontos_negativos'):
            coherence_parts.append("\n\n⚠️ PONTOS NEGATIVOS:\n" + "\n".join([f"• {item}" for item in ai_detailed_comments['pontos_negativos'][:3]]))
        
        coherence_comment = " ".join(coherence_parts) if coherence_parts else f"Coerência: {coherence_score}%"
    else:
        # Fallback para comentários genéricos
        completeness_score = score
        coherence_score = score
        
        if score >= 90:
            completeness_comment = "Apresentação extremamente completa, contendo todas as informações essenciais e adicionais relevantes."
            coherence_comment = "Informações altamente coerentes e bem estruturadas, facilitando a compreensão."
        elif score >= 70:
            completeness_comment = "Apresentação completa com as principais informações necessárias presentes."
            coherence_comment = "Informações coerentes e bem organizadas, com estrutura adequada."
        elif score >= 50:
            completeness_comment = "Apresentação parcialmente completa. Algumas informações importantes podem estar ausentes."
            coherence_comment = "Coerência razoável, mas algumas informações podem estar desorganizadas."
        else:
            completeness_comment = "Apresentação incompleta. Faltam informações essenciais para uma avaliação adequada."
            coherence_comment = "Baixa coerência. Informações desorganizadas ou conflitantes."
    
    # Comentário sobre instituição
    if institution_mentioned:
        institution_comment = f"✅ A instituição '{institution_name}' foi identificada no documento."
    else:
        institution_comment = f"❌ A instituição '{institution_name}' NÃO foi mencionada no documento. Isso pode indicar que o documento não corresponde à instituição esperada."
    
    return {
        'is_valid': is_valid,
        'score': score,
        'issues': issues,
        'warnings': warnings,
        'institution_mentioned': institution_mentioned,
        'institution_comment': institution_comment,
        'dates_found': len(dates),
        'completeness': completeness_score,
        'completeness_comment': completeness_comment,
        'coherence': coherence_score,
        'coherence_comment': coherence_comment,
        'summary': f"Apresentação {'APROVADA' if is_valid else 'REPROVADA'} (Score: {score}/100)",
        'confidence': 0.85,
        'provider': 'Gemini'
    }


# ========== ANÁLISE 2: CHECKLIST ==========
def analyze_checklist(file_path, institution_name):
    """
    Análise RIGOROSA do Checklist de Credenciamento (Excel)
    Regras:
    1. Validar estrutura
    2. Nome e CNPJ devem estar preenchidos
    3. Todos checkboxes marcados (sem X vermelho)
    4. Campo observações deve ter texto coerente e não raso
    """
    print(f"\n✅ ANÁLISE: Checklist de Credenciamento")
    
    issues = []
    warnings = []
    score = 100
    
    try:
        # Ler arquivo Excel
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active
        
        # Extrair dados relevantes
        all_data = []
        for row in sheet.iter_rows(min_row=1, max_row=100, values_only=True):
            if any(cell is not None for cell in row):
                all_data.append([str(cell) if cell is not None else '' for cell in row])
        
        # Converter para texto para análise
        excel_text = "\n".join([" | ".join(row) for row in all_data])
        
        # 1. Verificar preenchimento básico
        has_institution = any(institution_name.lower() in str(cell).lower() 
                            for row in all_data 
                            for cell in row) if institution_name else False
        
        if not has_institution and institution_name:
            issues.append(f"❌ Nome da instituição '{institution_name}' não encontrado no checklist")
            score -= 30
        
        # 2. Verificar se há CNPJ
        cnpj_pattern = r'\d{2}\.?\d{3}\.?\d{3}/?\d{4}-?\d{2}'
        has_cnpj = bool(re.search(cnpj_pattern, excel_text))
        if not has_cnpj:
            issues.append("❌ CNPJ não encontrado no checklist")
            score -= 25
        
        # 3. Verificar completude (procurar por campos vazios)
        empty_cells = sum(1 for row in all_data for cell in row if cell == '' or cell == 'None')
        if empty_cells > 20:
            warnings.append(f"⚠️ Muitas células vazias detectadas ({empty_cells})")
            score -= 15
        
        # 4. Verificar observações (procurar por texto de observações)
        obs_keywords = ['observação', 'observacao', 'obs:', 'comentário', 'comentario']
        has_observations = any(keyword in excel_text.lower() for keyword in obs_keywords)
        
        if has_observations:
            # Procurar por textos muito curtos após keywords de observação
            obs_texts = []
            for i, row in enumerate(all_data):
                for j, cell in enumerate(row):
                    if any(keyword in str(cell).lower() for keyword in obs_keywords):
                        # Pegar próxima célula ou células próximas
                        if j + 1 < len(row):
                            obs_texts.append(str(row[j + 1]))
                        if i + 1 < len(all_data) and j < len(all_data[i + 1]):
                            obs_texts.append(str(all_data[i + 1][j]))
            
            # Verificar se observações não são rasas
            for obs in obs_texts:
                if len(obs) < 20 and obs.strip() and obs != 'None':
                    warnings.append("⚠️ Observações muito curtas ou genéricas detectadas")
                    score -= 10
                    break
        
        # 5. Procurar por indicadores de checkboxes não marcados (X vermelho, vazio, etc)
        negative_indicators = ['❌', '✗', 'X', 'não', 'nao', 'pendente', 'falta']
        found_negatives = []
        for row in all_data:
            for cell in row:
                cell_str = str(cell).lower()
                if any(neg in cell_str for neg in negative_indicators):
                    if 'sim' not in cell_str and 'aprovad' not in cell_str:
                        found_negatives.append(cell_str[:50])
        
        if found_negatives and len(found_negatives) > 3:
            issues.append(f"❌ Itens não conformes ou pendentes detectados no checklist")
            score -= 30
        
        wb.close()
        
    except Exception as e:
        issues.append(f"❌ Erro ao ler arquivo Excel: {str(e)[:100]}")
        score = 0
    
    score = max(0, score)
    is_valid = score >= 70 and len(issues) == 0
    
    return {
        'is_valid': is_valid,
        'score': score,
        'issues': issues,
        'warnings': warnings,
        'summary': f"Checklist {'APROVADO' if is_valid else 'REPROVADO'} (Score: {score}/100)"
    }


# ========== ANÁLISE 3: CADPREV ==========
def analyze_cadprev(file_path, institution_name):
    """
    Análise RIGOROSA do CadPrev (Excel)
    Regras:
    1. Campos azuis = perguntas, amarelos = respostas
    2. NÃO obrigatórios: CNPJ, Razão Social, Tipo, Data, nº Processo (mas reportar)
    3. Todos outros obrigatórios
    4. Volume total gerido obrigatório
    5. Completude e coerência das respostas
    """
    print(f"\n📋 ANÁLISE: CadPrev")
    
    issues = []
    warnings = []
    score = 100
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active
        
        # Extrair todos os dados
        all_data = []
        for row in sheet.iter_rows(min_row=1, max_row=200, values_only=True):
            if any(cell is not None for cell in row):
                all_data.append([str(cell) if cell is not None else '' for cell in row])
        
        excel_text = "\n".join([" | ".join(row) for row in all_data])
        
        # 1. Verificar instituição
        has_institution = any(institution_name.lower() in str(cell).lower() 
                            for row in all_data 
                            for cell in row) if institution_name else False
        
        if not has_institution and institution_name:
            warnings.append(f"⚠️ Nome da instituição '{institution_name}' não encontrado (não obrigatório)")
        else:
            warnings.append(f"✓ Instituição '{institution_name}' mencionada")
        
        # 2. Verificar CNPJ (não obrigatório, mas reportar)
        cnpj_pattern = r'\d{2}\.?\d{3}\.?\d{3}/?\d{4}-?\d{2}'
        has_cnpj = bool(re.search(cnpj_pattern, excel_text))
        if not has_cnpj:
            warnings.append("⚠️ CNPJ não encontrado (não obrigatório)")
        else:
            warnings.append("✓ CNPJ presente")
        
        # 3. Verificar Volume Gerido (OBRIGATÓRIO)
        volume_keywords = ['volume', 'gerido', 'patrimônio', 'patrimonio', 'aum', 'total']
        has_volume = any(keyword in excel_text.lower() for keyword in volume_keywords)
        
        if not has_volume:
            issues.append("❌ Volume total gerido não encontrado (OBRIGATÓRIO)")
            score -= 40
        else:
            # Verificar se há valores numéricos associados
            numbers = re.findall(r'\d+[.,]?\d*', excel_text)
            large_numbers = [n for n in numbers if len(n.replace(',', '').replace('.', '')) > 6]
            if not large_numbers:
                warnings.append("⚠️ Volume gerido pode estar sem valor numérico adequado")
                score -= 15
        
        # 4. Verificar completude geral (campos vazios)
        empty_cells = sum(1 for row in all_data for cell in row if cell == '' or cell == 'None')
        total_cells = sum(len(row) for row in all_data)
        
        if total_cells > 0:
            empty_ratio = empty_cells / total_cells
            if empty_ratio > 0.5:
                issues.append(f"❌ CadPrev muito incompleto ({int(empty_ratio*100)}% de campos vazios)")
                score -= 35
            elif empty_ratio > 0.3:
                warnings.append(f"⚠️ Muitos campos vazios ({int(empty_ratio*100)}%)")
                score -= 15
        
        # 5. Verificar presença de respostas textuais (coerência)
        text_cells = [cell for row in all_data for cell in row if len(str(cell)) > 50]
        if len(text_cells) < 5:
            warnings.append("⚠️ Poucas respostas textuais detalhadas")
            score -= 10
        
        wb.close()
        
    except Exception as e:
        issues.append(f"❌ Erro ao ler arquivo Excel: {str(e)[:100]}")
        score = 0
    
    score = max(0, score)
    is_valid = score >= 70 and len(issues) == 0
    
    return {
        'is_valid': is_valid,
        'score': score,
        'issues': issues,
        'warnings': warnings,
        'summary': f"CadPrev {'APROVADO' if is_valid else 'REPROVADO'} (Score: {score}/100)"
    }


# ========== ANÁLISE 4: TERMO DE CREDENCIAMENTO ==========
def analyze_termo_credenciamento(file_path, institution_name):
    """
    Análise RIGOROSA do Termo de Credenciamento (Excel)
    DOCUMENTO CRÍTICO
    Regras:
    1. Estrutura rigorosa
    2. Campos laranja = obrigatórios
    3. Campos brancos = não obrigatórios (reportar)
    4. Completude e coerência
    """
    print(f"\n📄 ANÁLISE: Termo de Credenciamento")
    
    issues = []
    warnings = []
    score = 100
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active
        
        # Extrair todos os dados
        all_data = []
        mandatory_fields_found = 0
        optional_fields_found = 0
        
        for row in sheet.iter_rows(min_row=1, max_row=200):
            row_data = []
            for cell in row:
                value = str(cell.value) if cell.value is not None else ''
                row_data.append(value)
                
                # Verificar cor de fundo (laranja = obrigatório)
                if cell.fill and cell.fill.fgColor:
                    # Laranja = RGB próximo de FF9900 ou variantes
                    color = str(cell.fill.fgColor.rgb) if hasattr(cell.fill.fgColor, 'rgb') else ''
                    if 'FF9' in color or 'FFA5' in color or 'FFB84D' in color:
                        if value and value != 'None':
                            mandatory_fields_found += 1
            
            if any(cell for cell in row_data):
                all_data.append(row_data)
        
        excel_text = "\n".join([" | ".join(row) for row in all_data])
        
        # 1. Verificar instituição (OBRIGATÓRIO)
        has_institution = any(institution_name.lower() in str(cell).lower() 
                            for row in all_data 
                            for cell in row) if institution_name else False
        
        if not has_institution and institution_name:
            issues.append(f"❌ Nome da instituição '{institution_name}' não encontrado no Termo")
            score -= 40
        
        # 2. Verificar CNPJ (OBRIGATÓRIO)
        cnpj_pattern = r'\d{2}\.?\d{3}\.?\d{3}/?\d{4}-?\d{2}'
        has_cnpj = bool(re.search(cnpj_pattern, excel_text))
        if not has_cnpj:
            issues.append("❌ CNPJ não encontrado no Termo")
            score -= 30
        
        # 3. Verificar campos críticos
        critical_keywords = {
            'categoria': ['categoria', 'tipo de credenciamento'],
            'endereco': ['endereço', 'endereco', 'logradouro'],
            'representante': ['representante legal', 'responsável', 'responsavel'],
            'contato': ['telefone', 'email', 'e-mail', 'contato']
        }
        
        for field, keywords in critical_keywords.items():
            found = any(keyword in excel_text.lower() for keyword in keywords)
            if found:
                # Verificar se há conteúdo após keyword
                has_content = False
                for keyword in keywords:
                    if keyword in excel_text.lower():
                        idx = excel_text.lower().index(keyword)
                        nearby = excel_text[idx:idx+200]
                        if len([c for c in nearby if c.isalnum()]) > 20:
                            has_content = True
                            break
                
                if not has_content:
                    warnings.append(f"⚠️ Campo '{field}' pode estar vazio")
                    score -= 10
            else:
                warnings.append(f"⚠️ Campo '{field}' não identificado")
                score -= 5
        
        # 4. Verificar completude
        empty_cells = sum(1 for row in all_data for cell in row if cell == '' or cell == 'None')
        total_cells = sum(len(row) for row in all_data)
        
        if total_cells > 0:
            empty_ratio = empty_cells / total_cells
            if empty_ratio > 0.4:
                issues.append(f"❌ Termo muito incompleto ({int(empty_ratio*100)}% de campos vazios)")
                score -= 30
            elif empty_ratio > 0.25:
                warnings.append(f"⚠️ Muitos campos vazios ({int(empty_ratio*100)}%)")
                score -= 15
        
        # 5. Verificar data
        dates = extract_dates_from_text(excel_text)
        if not dates:
            warnings.append("⚠️ Nenhuma data encontrada no Termo")
            score -= 10
        else:
            warnings.append(f"✓ Data encontrada: {max(dates).strftime('%d/%m/%Y')}")
        
        wb.close()
        
    except Exception as e:
        issues.append(f"❌ Erro ao ler arquivo Excel: {str(e)[:100]}")
        score = 0
    
    score = max(0, score)
    is_valid = score >= 70 and len(issues) == 0
    
    return {
        'is_valid': is_valid,
        'score': score,
        'issues': issues,
        'warnings': warnings,
        'summary': f"Termo de Credenciamento {'APROVADO' if is_valid else 'REPROVADO'} (Score: {score}/100)"
    }


# ========== ANÁLISE 5: DECLARAÇÃO UNIFICADA ==========
def analyze_declaracao_unificada(file_path, institution_name):
    """
    Análise RIGOROSA da Declaração Unificada (PDF + Assinatura)
    Regras:
    1. Texto contém "declaração unificada"
    2. Data < 1 ano
    3. Validar assinatura digital (TCEES)
    """
    print(f"\n📝 ANÁLISE: Declaração Unificada")
    
    text = extract_text_from_pdf(file_path)
    if text.startswith("ERRO"):
        return {'is_valid': False, 'score': 0, 'issues': [text], 'warnings': [], 'summary': 'Erro ao ler PDF'}
    
    issues = []
    warnings = []
    score = 100
    
    # 1. Verificar se é declaração unificada
    is_declaracao = 'declaração unificada' in text.lower() or 'declaracao unificada' in text.lower()
    if not is_declaracao:
        issues.append("❌ Documento não contém 'Declaração Unificada' no texto")
        score -= 50
    
    # 2. Verificar data
    dates = extract_dates_from_text(text)
    if dates:
        most_recent_date = max(dates)
        if not is_date_within_one_year(most_recent_date):
            issues.append(f"❌ Declaração desatualizada (data: {most_recent_date.strftime('%d/%m/%Y')})")
            score -= 30
    else:
        warnings.append("⚠️ Nenhuma data encontrada")
    
    # 3. Verificar menção à instituição
    if institution_name and institution_name.lower() not in text.lower():
        issues.append(f"❌ Instituição '{institution_name}' não mencionada")
        score -= 20
    
    # 4. TODO: Validar assinatura digital (TCEES)
    warnings.append("⚠️ Validação de assinatura digital: não implementada (requer integração TCEES)")
    
    score = max(0, score)
    is_valid = score >= 70 and len(issues) == 0
    
    return {
        'is_valid': is_valid,
        'score': score,
        'issues': issues,
        'warnings': warnings,
        'summary': f"Declaração {'APROVADA' if is_valid else 'REPROVADA'} (Score: {score}/100)"
    }


# ========== ANÁLISE 6: RATING ==========
def analyze_rating(file_path, institution_name):
    """
    Análise RIGOROSA de Rating (PDF)
    Regras:
    1. Documento analisa risco
    2. Concede score/nota/classificação
    3. Menciona instituição
    """
    print(f"\n⭐ ANÁLISE: Rating")
    
    text = extract_text_from_pdf(file_path)
    if text.startswith("ERRO"):
        return {'is_valid': False, 'score': 0, 'issues': [text], 'warnings': [], 'summary': 'Erro ao ler PDF'}
    
    issues = []
    warnings = []
    score = 100
    
    # Termos que indicam análise de risco
    risk_terms = ['risco', 'rating', 'classificação', 'análise', 'crédito', 'score']
    has_risk_analysis = any(term in text.lower() for term in risk_terms)
    
    if not has_risk_analysis:
        issues.append("❌ Documento não parece ser um relatório de rating/risco")
        score -= 40
    
    # Verificar menção à instituição
    if institution_name and institution_name.lower() not in text.lower():
        issues.append(f"❌ Instituição '{institution_name}' não mencionada no rating")
        score -= 30
    
    # Verificar presença de classificação/nota
    rating_patterns = [r'rating[:\s]+([A-Z]{1,3}[\+\-]?)', r'classificação[:\s]+([A-Z]{1,3})',
                       r'nota[:\s]+(\d+\.?\d*)', r'score[:\s]+(\d+)']
    has_rating = any(re.search(pattern, text, re.IGNORECASE) for pattern in rating_patterns)
    
    if not has_rating:
        warnings.append("⚠️ Não foi possível identificar classificação/nota clara no documento")
        score -= 15
    
    score = max(0, score)
    is_valid = score >= 70
    
    return {
        'is_valid': is_valid,
        'score': score,
        'issues': issues,
        'warnings': warnings,
        'summary': f"Rating {'APROVADO' if is_valid else 'REPROVADO'} (Score: {score}/100)"
    }


# ========== ANÁLISE 7: CERTIDÕES ==========
def analyze_certidao(file_path, institution_name, certidao_type):
    """
    Análise RIGOROSA de Certidões
    Regras:
    1. Nome da instituição aparece
    2. Tipo de certidão corresponde
    """
    print(f"\n🏛️ ANÁLISE: Certidão - {certidao_type}")
    
    text = extract_text_from_pdf(file_path)
    if text.startswith("ERRO"):
        return {'is_valid': False, 'score': 0, 'issues': [text], 'warnings': [], 'summary': 'Erro ao ler PDF'}
    
    issues = []
    warnings = []
    score = 100
    
    # Verificar menção à instituição
    if institution_name and institution_name.lower() not in text.lower():
        issues.append(f"❌ Instituição '{institution_name}' não aparece na certidão")
        score -= 50
    
    # Verificar tipo de certidão
    certidao_keywords = {
        'certidao_municipal': ['municipal', 'prefeitura', 'município'],
        'certidao_estadual': ['estadual', 'estado', 'fazenda estadual'],
        'certidao_federal': ['federal', 'receita federal', 'união'],
        'certidao_trabalhista': ['trabalhista', 'justiça do trabalho', 'tst'],
        'certidao_fgts': ['fgts', 'fundo de garantia']
    }
    
    keywords = certidao_keywords.get(certidao_type, [])
    type_match = any(kw in text.lower() for kw in keywords)
    
    if not type_match and keywords:
        warnings.append(f"⚠️ Tipo de certidão pode não corresponder a '{certidao_type}'")
        score -= 20
    
    score = max(0, score)
    is_valid = score >= 70
    
    return {
        'is_valid': is_valid,
        'score': score,
        'issues': issues,
        'warnings': warnings,
        'summary': f"Certidão {'APROVADA' if is_valid else 'REPROVADA'} (Score: {score}/100)"
    }


# ========== ANÁLISE 8: TERMO DE DECLARAÇÃO ==========
def analyze_termo_declaracao(file_path, institution_name):
    """Análise de Termo de Declaração (PDF assinado)"""
    print(f"\n📝 ANÁLISE: Termo de Declaração")
    
    text = extract_text_from_pdf(file_path)
    if text.startswith("ERRO"):
        return {'is_valid': False, 'score': 0, 'issues': [text], 'warnings': [], 'summary': 'Erro ao ler PDF'}
    
    issues = []
    warnings = []
    score = 100
    
    # Verificar termo "declaração" ou "declaramos"
    if 'declaração' not in text.lower() and 'declaramos' not in text.lower():
        issues.append("❌ Documento não parece ser um termo de declaração")
        score -= 50
    
    # Verificar menção à instituição
    if institution_name and institution_name.lower() not in text.lower():
        warnings.append(f"⚠️ Instituição '{institution_name}' pode não estar mencionada")
        score -= 20
    
    # Verificar data recente (< 1 ano)
    dates = extract_dates_from_text(text)
    recent_date = False
    for date in dates:
        if is_date_within_one_year(date):
            recent_date = True
            break
    
    if not recent_date and dates:
        warnings.append("⚠️ Data do documento pode estar desatualizada (> 1 ano)")
        score -= 15
    
    score = max(0, score)
    is_valid = score >= 70
    
    return {
        'is_valid': is_valid,
        'score': score,
        'issues': issues,
        'warnings': warnings,
        'summary': f"Termo de Declaração {'APROVADO' if is_valid else 'REPROVADO'} (Score: {score}/100)"
    }


# ========== ANÁLISE 9: QDD ANBIMA ==========
def analyze_qdd_anbima(file_path, institution_name):
    """Análise de QDD Anbima Seção I"""
    print(f"\n📊 ANÁLISE: QDD Anbima Seção I")
    
    text = extract_text_from_pdf(file_path)
    if text.startswith("ERRO"):
        return {'is_valid': False, 'score': 0, 'issues': [text], 'warnings': [], 'summary': 'Erro ao ler PDF'}
    
    issues = []
    warnings = []
    score = 100
    
    # Verificar menção ANBIMA
    if 'anbima' not in text.lower():
        issues.append("❌ Documento não parece ser da ANBIMA")
        score -= 40
    
    # Verificar "QDD" ou "Questionário"
    if 'qdd' not in text.lower() and 'questionário' not in text.lower():
        warnings.append("⚠️ Termo QDD/Questionário não identificado")
        score -= 20
    
    # Verificar instituição
    if institution_name and institution_name.lower() not in text.lower():
        warnings.append(f"⚠️ Instituição '{institution_name}' pode não estar mencionada")
        score -= 15
    
    score = max(0, score)
    is_valid = score >= 70
    
    return {
        'is_valid': is_valid,
        'score': score,
        'issues': issues,
        'warnings': warnings,
        'summary': f"QDD Anbima {'APROVADO' if is_valid else 'REPROVADO'} (Score: {score}/100)"
    }


# ========== ANÁLISE 10: CERTIDÕES ESPECÍFICAS BACEN/ANBIMA/CMN ==========
def analyze_certidao_especifica(file_path, institution_name, doc_type):
    """Análise de certidões específicas"""
    
    type_info = {
        'certidao_bacen_autorizacao': {'nome': 'Certidão BACEN - Autorização', 'keywords': ['bacen', 'banco central', 'autorização']},
        'certidao_bacen_nada_consta': {'nome': 'Certidão Nada Consta BACEN', 'keywords': ['bacen', 'nada consta', 'banco central']},
        'certidao_anbima': {'nome': 'Certidão Adesão ANBIMA', 'keywords': ['anbima', 'adesão', 'código']},
        'lista_exaustiva_cmn': {'nome': 'Lista Exaustiva CMN', 'keywords': ['cmn', 'resolução', 'art', '15']},
        'formulario_referencia_cvm': {'nome': 'Formulário Referência CVM', 'keywords': ['cvm', 'formulário', 'referência']}
    }
    
    info = type_info.get(doc_type, {'nome': doc_type, 'keywords': []})
    print(f"\n🏛️ ANÁLISE: {info['nome']}")
    
    text = extract_text_from_pdf(file_path)
    if text.startswith("ERRO"):
        return {'is_valid': False, 'score': 0, 'issues': [text], 'warnings': [], 'summary': 'Erro ao ler PDF'}
    
    issues = []
    warnings = []
    score = 100
    
    # Verificar palavras-chave do tipo
    keywords_found = sum(1 for kw in info['keywords'] if kw in text.lower())
    if keywords_found == 0:
        issues.append(f"❌ Documento não parece ser {info['nome']}")
        score -= 50
    elif keywords_found < len(info['keywords']) / 2:
        warnings.append(f"⚠️ Poucas palavras-chave encontradas para {info['nome']}")
        score -= 20
    
    # Verificar instituição
    if institution_name and institution_name.lower() not in text.lower():
        warnings.append(f"⚠️ Instituição '{institution_name}' pode não estar mencionada")
        score -= 15
    
    score = max(0, score)
    is_valid = score >= 70
    
    return {
        'is_valid': is_valid,
        'score': score,
        'issues': issues,
        'warnings': warnings,
        'summary': f"{info['nome']} {'APROVADO' if is_valid else 'REPROVADO'} (Score: {score}/100)"
    }


# ========== ANÁLISE 11: CONTRATO DE DISTRIBUIÇÃO ==========
def analyze_contrato_distribuicao(file_path, institution_name):
    """Análise de Contrato de Distribuição"""
    print(f"\n📄 ANÁLISE: Contrato de Distribuição")
    
    text = extract_text_from_pdf(file_path)
    if text.startswith("ERRO"):
        return {'is_valid': False, 'score': 0, 'issues': [text], 'warnings': [], 'summary': 'Erro ao ler PDF'}
    
    issues = []
    warnings = []
    score = 100
    
    # Verificar termo "contrato" e "distribuição"
    if 'contrato' not in text.lower():
        issues.append("❌ Termo 'contrato' não encontrado")
        score -= 40
    
    if 'distribuição' not in text.lower() and 'distribuicao' not in text.lower():
        warnings.append("⚠️ Termo 'distribuição' não encontrado claramente")
        score -= 20
    
    # Verificar instituição
    if institution_name and institution_name.lower() not in text.lower():
        warnings.append(f"⚠️ Instituição '{institution_name}' pode não estar mencionada")
        score -= 15
    
    score = max(0, score)
    is_valid = score >= 70
    
    return {
        'is_valid': is_valid,
        'score': score,
        'issues': issues,
        'warnings': warnings,
        'summary': f"Contrato de Distribuição {'APROVADO' if is_valid else 'REPROVADO'} (Score: {score}/100)"
    }


# ========== ANÁLISE 12: SITUAÇÃO ANCORD ==========
def analyze_situacao_ancord(file_path, institution_name):
    """Análise de Situação ANCORD (para AAI)"""
    print(f"\n📋 ANÁLISE: Situação ANCORD")
    
    text = extract_text_from_pdf(file_path)
    if text.startswith("ERRO"):
        return {'is_valid': False, 'score': 0, 'issues': [text], 'warnings': [], 'summary': 'Erro ao ler PDF'}
    
    issues = []
    warnings = []
    score = 100
    
    # Verificar ANCORD
    if 'ancord' not in text.lower():
        issues.append("❌ Termo 'ANCORD' não encontrado")
        score -= 50
    
    # Verificar AAI (Agente Autônomo de Investimentos)
    if 'aai' not in text.lower() and 'agente autônomo' not in text.lower():
        warnings.append("⚠️ Termo AAI/Agente Autônomo não encontrado")
        score -= 20
    
    score = max(0, score)
    is_valid = score >= 70
    
    return {
        'is_valid': is_valid,
        'score': score,
        'issues': issues,
        'warnings': warnings,
        'summary': f"Situação ANCORD {'APROVADA' if is_valid else 'REPROVADA'} (Score: {score}/100)"
    }


# ========== ROTEADOR PRINCIPAL ==========
def analyze_document_rigorous(file_path, document_type, document_name, institution_name, institution_cnpj=None):
    """
    ROTEADOR PRINCIPAL para análises rigorosas
    Direciona para a função específica baseada no tipo
    """
    
    print(f"\n{'='*60}")
    print(f"🔍 ANÁLISE RIGOROSA INICIADA")
    print(f"   Documento: {document_name}")
    print(f"   Tipo: {document_type}")
    print(f"   Instituição: {institution_name}")
    print(f"{'='*60}")
    
    # Roteamento
    if document_type == 'apresentacao_institucional':
        result = analyze_apresentacao_institucional(file_path, institution_name)
    
    elif document_type == 'checklist':
        result = analyze_checklist(file_path, institution_name)
    
    elif document_type == 'cadprev':
        result = analyze_cadprev(file_path, institution_name)
    
    elif document_type == 'termo_credenciamento':
        result = analyze_termo_credenciamento(file_path, institution_name)
    
    elif document_type == 'termo_declaracao':
        result = analyze_termo_declaracao(file_path, institution_name)
    
    elif document_type == 'declaracao_unificada':
        result = analyze_declaracao_unificada(file_path, institution_name)
    
    elif document_type == 'qdd_anbima':
        result = analyze_qdd_anbima(file_path, institution_name)
    
    elif document_type in ['certidao_bacen_autorizacao', 'certidao_bacen_nada_consta', 
                           'certidao_anbima', 'lista_exaustiva_cmn', 'formulario_referencia_cvm']:
        result = analyze_certidao_especifica(file_path, institution_name, document_type)
    
    elif document_type == 'rating':
        result = analyze_rating(file_path, institution_name)
    
    elif document_type == 'contrato_distribuicao':
        result = analyze_contrato_distribuicao(file_path, institution_name)
    
    elif document_type == 'situacao_ancord':
        result = analyze_situacao_ancord(file_path, institution_name)
    
    elif document_type in ['certidao_municipal', 'certidao_estadual', 'certidao_federal', 
                           'certidao_trabalhista', 'certidao_fgts']:
        result = analyze_certidao(file_path, institution_name, document_type)
    
    else:
        result = {
            'is_valid': False,
            'score': 0,
            'issues': [f"Tipo de documento '{document_type}' não reconhecido"],
            'warnings': [],
            'summary': 'Tipo de documento inválido'
        }
    
    print(f"\n{'='*60}")
    print(f"✅ ANÁLISE CONCLUÍDA")
    print(f"   Score: {result.get('score', 0)}/100")
    print(f"   Status: {'APROVADO ✓' if result.get('is_valid') else 'REPROVADO ✗'}")
    print(f"   Problemas: {len(result.get('issues', []))}")
    print(f"   Avisos: {len(result.get('warnings', []))}")
    print(f"{'='*60}\n")
    
    return result

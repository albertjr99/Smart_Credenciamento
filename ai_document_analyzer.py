"""
Módulo de Análise Inteligente de Documentos com IA
Sistema Profissional e Robusto para Produção
Suporta múltiplos provedores de IA com fallback automático

INTEGRADO COM:
- Base de Conhecimento do Agente (ai_document_knowledge.py)
- Validador de Assinatura Digital TCEES (tcees_validator.py)
"""

import PyPDF2
import openpyxl
from openpyxl.styles import PatternFill
import re
from datetime import datetime, timedelta
from dateutil import parser as date_parser
import json
import os

# Importar motor de IA robusto
from ai_config import get_ai_analysis, get_ai_status

# Importar base de conhecimento do agente
from ai_document_knowledge import (
    get_document_knowledge,
    get_ai_prompt_for_document,
    get_validation_rules,
    is_signature_critical,
    get_signature_action,
    check_excel_cell_color,
    is_check_symbol,
    is_x_symbol,
    EXCEL_COLORS,
    CHECK_SYMBOLS,
    X_SYMBOLS,
    SIGNATURE_RULES
)

# Importar validador TCEES
try:
    from tcees_validator import validate_pdf_with_tcees
    TCEES_AVAILABLE = True
except ImportError:
    TCEES_AVAILABLE = False
    print("⚠️ Validador TCEES não disponível")


# =============================================================================
# FUNÇÕES DE ANÁLISE EXCEL INTELIGENTE
# =============================================================================

def get_cell_background_color(cell):
    """
    Extrai a cor de fundo de uma célula do Excel
    Retorna o código RGB ou None
    """
    try:
        if cell.fill and cell.fill.patternType:
            fg_color = cell.fill.fgColor
            if fg_color:
                if fg_color.type == 'rgb' and fg_color.rgb:
                    return str(fg_color.rgb)
                elif fg_color.type == 'indexed':
                    # Índices de cores padrão do Excel
                    indexed_colors = {
                        0: 'FF000000', 1: 'FFFFFFFF', 2: 'FFFF0000', 3: 'FF00FF00',
                        4: 'FF0000FF', 5: 'FFFFFF00', 6: 'FFFF00FF', 7: 'FF00FFFF',
                        10: 'FF00FF00', 11: 'FF0000FF', 13: 'FFFFFF00', 22: 'FFC0C0C0',
                        44: 'FFFFFF99', 45: 'FF99CCFF', 50: 'FFFF6600', 51: 'FF99CC00'
                    }
                    return indexed_colors.get(fg_color.indexed, None)
        return None
    except:
        return None


def analyze_excel_by_color(file_path, color_rules):
    """
    Analisa uma planilha Excel baseado em regras de cores
    
    Args:
        file_path: Caminho para o arquivo Excel
        color_rules: Dict com regras de cores {'color_type': 'validation_rule'}
    
    Returns:
        Dict com resultados da análise por cor
    """
    result = {
        'cells_by_color': {},
        'filled_by_color': {},
        'empty_by_color': {},
        'content_by_color': {},
        'issues': [],
        'warnings': []
    }
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active
        
        for row in sheet.iter_rows():
            for cell in row:
                cell_color = get_cell_background_color(cell)
                if not cell_color:
                    continue
                
                cell_value = cell.value
                is_filled = cell_value is not None and str(cell_value).strip() != ''
                
                # Verificar cada tipo de cor definido
                for color_type, color_values in EXCEL_COLORS.items():
                    if any(cv.upper() in cell_color.upper() for cv in color_values):
                        # Inicializar contadores
                        if color_type not in result['cells_by_color']:
                            result['cells_by_color'][color_type] = 0
                            result['filled_by_color'][color_type] = 0
                            result['empty_by_color'][color_type] = 0
                            result['content_by_color'][color_type] = []
                        
                        result['cells_by_color'][color_type] += 1
                        
                        if is_filled:
                            result['filled_by_color'][color_type] += 1
                            result['content_by_color'][color_type].append(str(cell_value))
                        else:
                            result['empty_by_color'][color_type] += 1
                        
                        break
        
        wb.close()
        
    except Exception as e:
        result['issues'].append(f"Erro ao analisar Excel: {str(e)}")
    
    return result


def extract_excel_content_for_ai(file_path, max_chars=6000):
    """
    Extrai conteúdo do Excel em formato texto para análise de IA
    Preserva estrutura e informações de cores
    """
    content = []
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active
        
        content.append(f"=== PLANILHA: {sheet.title} ===\n")
        content.append(f"Dimensões: {sheet.max_row} linhas x {sheet.max_column} colunas\n")
        
        for row_idx, row in enumerate(sheet.iter_rows(), 1):
            row_content = []
            for cell in row:
                if cell.value:
                    cell_color = get_cell_background_color(cell)
                    color_info = ""
                    
                    # Identificar cor significativa
                    if cell_color:
                        for color_type, color_values in EXCEL_COLORS.items():
                            if any(cv.upper() in cell_color.upper() for cv in color_values):
                                if 'green' in color_type:
                                    color_info = "[✓]"
                                elif 'red' in color_type:
                                    color_info = "[✗]"
                                elif 'yellow' in color_type or 'answer' in color_type:
                                    color_info = "[RESPOSTA]"
                                elif 'orange' in color_type or 'institution' in color_type:
                                    color_info = "[OBRIGATÓRIO]"
                                elif 'blue' in color_type or 'question' in color_type:
                                    color_info = "[PERGUNTA]"
                                break
                    
                    row_content.append(f"{color_info}{cell.value}")
            
            if row_content:
                content.append(f"L{row_idx}: " + " | ".join(row_content))
            
            # Limitar tamanho
            if len("\n".join(content)) > max_chars:
                content.append("\n... [conteúdo truncado] ...")
                break
        
        wb.close()
        
    except Exception as e:
        content.append(f"Erro ao extrair Excel: {str(e)}")
    
    return "\n".join(content)


def validate_signature_if_applicable(file_path, document_type):
    """
    Valida assinatura digital se o documento for PDF e a validação estiver disponível
    
    Returns:
        Dict com resultado da validação ou None se não aplicável
    """
    if not TCEES_AVAILABLE:
        return {
            'validated': False,
            'reason': 'Validador TCEES não disponível',
            'is_critical': is_signature_critical(document_type)
        }
    
    if not file_path.lower().endswith('.pdf'):
        return None
    
    try:
        tcees_result = validate_pdf_with_tcees(file_path)
        
        is_signed = tcees_result.get('assinado', False)
        is_valid = tcees_result.get('autenticidade_ok', False) and tcees_result.get('integridade_ok', False)
        
        signature_action = get_signature_action(document_type, is_valid if is_signed else False)
        
        return {
            'validated': True,
            'is_signed': is_signed,
            'signature_valid': is_valid,
            'is_critical': is_signature_critical(document_type),
            'action': signature_action['action'],
            'message': signature_action['message'],
            'tcees_result': tcees_result
        }
        
    except Exception as e:
        return {
            'validated': False,
            'reason': f'Erro na validação: {str(e)}',
            'is_critical': is_signature_critical(document_type)
        }


def extract_text_from_pdf(file_path):
    """Extrai texto de um arquivo PDF"""
    try:
        with open(file_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            text = ""
            for page in pdf_reader.pages:
                text += page.extract_text() + "\n"
            return text
    except Exception as e:
        return f"Erro ao extrair texto: {str(e)}"

def extract_dates_from_text(text):
    """Extrai datas do texto"""
    dates = []
    # Padrões de data comuns
    patterns = [
        r'\d{1,2}[/-]\d{1,2}[/-]\d{2,4}',  # DD/MM/YYYY ou DD-MM-YYYY
        r'\d{1,2}\s+de\s+\w+\s+de\s+\d{4}',  # DD de Mês de YYYY
        r'\w+\s+\d{1,2},\s+\d{4}',  # Mês DD, YYYY
    ]
    
    for pattern in patterns:
        matches = re.findall(pattern, text, re.IGNORECASE)
        for match in matches:
            try:
                # Tentar parsear a data
                date_obj = date_parser.parse(match, fuzzy=True, dayfirst=True)
                dates.append(date_obj)
            except:
                continue
    
    return dates

def is_date_within_one_year(date_obj):
    """Verifica se a data está dentro do último ano"""
    if not date_obj:
        return None
    
    today = datetime.now()
    one_year_ago = today - timedelta(days=365)
    return date_obj >= one_year_ago

def analyze_with_advanced_ai(prompt, context="", document_type=""):
    """
    Usa sistema robusto de IA multi-provedor para análise profunda
    Suporta OpenAI GPT-4, Anthropic Claude, Google Gemini
    """
    
    # Verificar status da IA
    ai_status = get_ai_status()
    
    if not ai_status['available']:
        return {
            'success': False,
            'error': f'IA não configurada. Provedores disponíveis: {ai_status["all_providers"]}',
            'fallback': True
        }
    
    try:
        result = get_ai_analysis(prompt, context, document_type)
        
        if result['success']:
            return {
                'success': True,
                'analysis': result['analysis'],
                'provider': result.get('provider', 'Desconhecido'),
                'tokens_used': result.get('tokens_used', 0),
                'cost_estimate': result.get('cost_estimate', 0),
                'engine_info': result.get('engine_info', {}),
                'ai_powered': True
            }
        else:
            return {
                'success': False,
                'error': result.get('error', 'Erro desconhecido'),
                'fallback': True
            }
            
    except Exception as e:
        print(f"Erro ao usar IA avançada: {e}")
        return {
            'success': False,
            'error': str(e),
            'fallback': True
        }

# ============================================
# 1. ANÁLISE: APRESENTAÇÃO INSTITUCIONAL
# ============================================
def analyze_apresentacao_institucional(file_path, institution_name):
    """
    Analisa apresentação institucional (PDF ou PPTX)
    - Verifica se fala sobre a instituição
    - Verifica data (não pode ter mais de 1 ano)
    - Análise de conteúdo com IA
    """
    print(f"\n📊 analyze_apresentacao_institucional() iniciada")
    print(f"   - Arquivo: {file_path}")
    print(f"   - Instituição: {institution_name}")
    
    result = {
        'document_type': 'Apresentação Institucional',
        'is_valid': False,
        'issues': [],
        'warnings': [],
        'details': {},
        'score': 0
    }
    
    try:
        # Extrair texto
        print(f"   📄 Extraindo texto do PDF...")
        text = extract_text_from_pdf(file_path)
        print(f"   ✓ Texto extraído: {len(text)} caracteres")
        
        if not text or len(text.strip()) < 100:
            result['issues'].append("❌ Documento vazio ou com muito pouco conteúdo (menos de 100 caracteres)")
            print(f"   ❌ Documento rejeitado: conteúdo insuficiente")
            return result
        
        # ANÁLISE COM IA AVANÇADA (PRIORIDADE MÁXIMA)
        print(f"   🤖 Chamando IA para análise avançada...")
        gpt_prompt = f"""Analise esta Apresentação Institucional com RIGOR PROFISSIONAL:

INSTITUIÇÃO ESPERADA: "{institution_name}"
TIPO DE DOCUMENTO ESPERADO: Apresentação Institucional

VERIFICAÇÕES CRÍTICAS OBRIGATÓRIAS:
1. Este documento é REALMENTE uma apresentação institucional? 
   - Deve conter: visão geral da empresa, histórico, serviços, diferenciais, equipe, casos de sucesso
   - NÃO pode ser: termo, checklist, certidão, declaração, relatório financeiro, contrato

2. O documento menciona explicitamente a instituição "{institution_name}"?
   - Busque o nome completo ou partes significativas do nome
   - Verifique se fala SOBRE esta instituição (não apenas menciona)

3. Qualidade e completude do conteúdo:
   - Informações são suficientes e relevantes para credenciamento?
   - Conteúdo é específico da empresa ou genérico/copiado?
   - Há dados institucionais relevantes (experiência, AUM, equipe, etc)?

4. Coerência para credenciamento RPPS:
   - Conteúdo é adequado para análise de credenciamento?
   - Há informações sobre governança, compliance, processos?

REJEITE AUTOMATICAMENTE SE:
- Documento não é apresentação institucional (é outro tipo)
- Não menciona a instituição correta
- Conteúdo raso, genérico ou irrelevante
- Trata de outra empresa ou assunto
- Informações insuficientes para análise de credenciamento

SEJA EXTREMAMENTE CRÍTICO. Este é um sistema comercial."""

        ai_result = analyze_with_advanced_ai(gpt_prompt, text, "Apresentação Institucional")
        
        if ai_result.get('success') and ai_result.get('ai_powered'):
            # Usar análise da IA avançada
            ai_analysis = ai_result['analysis']
            
            result['is_valid'] = ai_analysis.get('is_valid', False)
            result['score'] = ai_analysis.get('score', 0)
            result['issues'] = ai_analysis.get('issues', [])
            result['warnings'] = ai_analysis.get('warnings', [])
            
            result['details'] = {
                'ai_powered': True,
                'provider': ai_result.get('provider', 'Desconhecido'),
                'confidence': ai_analysis.get('confidence_score', 0),
                'document_type_correct': ai_analysis.get('document_type_correct', False),
                'institution_mentioned': ai_analysis.get('institution_mentioned', False),
                'content_quality': ai_analysis.get('content_quality', 'unknown'),
                'completeness': ai_analysis.get('completeness', 0),
                'coherence': ai_analysis.get('coherence', 0),
                'summary': ai_analysis.get('summary', ''),
                'extracted_data': ai_analysis.get('extracted_data', {}),
                'tokens_used': ai_result.get('tokens_used', 0),
                'cost_usd': ai_result.get('cost_estimate', 0)
            }
            
            # Adicionar recomendações se houver
            if 'recommendations' in ai_analysis:
                result['recommendations'] = ai_analysis['recommendations']
            
            # Adicionar análise detalhada
            if 'detailed_analysis' in ai_analysis:
                result['details']['detailed_analysis'] = ai_analysis['detailed_analysis']
                
        else:
            # Fallback: análise baseada em regras MELHORADA
            result['warnings'].append("⚠️ Análise avançada com IA não disponível - usando análise básica")
            result['details']['ai_powered'] = False
            
            text_lower = text.lower()
            
            # VERIFICAÇÃO CRÍTICA: Detectar se é outro tipo de documento
            wrong_doc_indicators = {
                'termo de credenciamento': ['termo de análise', 'termo de credenciamento', 'anexo i', 'checklist'],
                'checklist': ['checklist', 'requisito', 'atribuição', 'situação'],
                'certidão': ['certidão', 'certifica', 'certificado'],
                'declaração': ['declara que', 'declaração', 'pelo presente instrumento'],
                'formulário': ['formulário', 'preencher', 'campo obrigatório'],
                'relatório financeiro': ['balanço', 'patrimônio líquido', 'demonstrativo', 'ativo', 'passivo']
            }
            
            # Verificar se parece ser outro tipo de documento
            for doc_type, indicators in wrong_doc_indicators.items():
                matches = sum(1 for indicator in indicators if indicator in text_lower)
                if matches >= 2:  # Se encontrar 2+ indicadores de outro tipo
                    result['issues'].append(f"❌ ERRO: Documento parece ser '{doc_type}', não uma apresentação institucional")
                    result['score'] = 0
                    result['is_valid'] = False
                    return result
            
            # Verificar se parece apresentação institucional
            presentation_indicators = [
                'apresentação', 'sobre nós', 'nossa história', 'quem somos',
                'missão', 'visão', 'valores', 'serviços', 'produtos',
                'experiência', 'equipe', 'clientes', 'portfólio', 'atuação'
            ]
            
            presentation_score = sum(1 for indicator in presentation_indicators if indicator in text_lower)
            
            if presentation_score == 0:
                result['issues'].append("❌ Documento NÃO parece ser uma apresentação institucional (sem indicadores típicos)")
                result['score'] = 0
                result['is_valid'] = False
                return result
            elif presentation_score < 3:
                result['warnings'].append(f"⚠️ Poucos indicadores de apresentação institucional encontrados ({presentation_score})")
                result['score'] = 30
            else:
                result['score'] = min(50, presentation_score * 10)
            
            # Verificar se menciona a instituição
            institution_mentioned = False
            if institution_name:
                name_parts = institution_name.lower().split()
                for part in name_parts:
                    if len(part) > 3 and part in text_lower:
                        institution_mentioned = True
                        break
            
            if not institution_mentioned:
                result['issues'].append(f"❌ A apresentação NÃO menciona a instituição '{institution_name}'")
                result['score'] = max(0, result['score'] - 30)
            else:
                result['details']['institution_mentioned'] = True
                result['score'] += 30
        
        # Verificar data (independente de GPT)
        dates = extract_dates_from_text(text)
        
        if dates:
            most_recent_date = max(dates)
            result['details']['date_found'] = most_recent_date.strftime('%d/%m/%Y')
            
            if is_date_within_one_year(most_recent_date):
                result['details']['date_valid'] = True
                if not result.get('is_valid'):  # Se GPT não validou, adicionar pontos
                    result['score'] += 50
            else:
                result['issues'].append(f"❌ Data da apresentação ({most_recent_date.strftime('%d/%m/%Y')}) é muito antiga (mais de 1 ano)")
                result['details']['date_valid'] = False
                result['is_valid'] = False
                result['score'] = max(0, result['score'] - 30)
        else:
            result['warnings'].append("⚠️ Não foi possível identificar a data da apresentação no documento")
            result['details']['date_found'] = None
        
        # Decisão final: documento só é válido se não houver issues críticos
        if len(result['issues']) > 0:
            result['is_valid'] = False
            result['score'] = min(result['score'], 40)  # Score máximo 40 com issues
        elif result['score'] < 50:
            result['is_valid'] = False
        
    except Exception as e:
        result['issues'].append(f"❌ Erro ao analisar documento: {str(e)}")
        result['is_valid'] = False
    
    return result

# ============================================
# 2. ANÁLISE: CHECKLIST CREDENCIAMENTO (EXCEL)
# ============================================
def analyze_checklist_credenciamento(file_path, institution_name, institution_cnpj):
    """
    Analisa checklist de credenciamento (Excel) com CONHECIMENTO ESPECIALIZADO
    
    REGRAS DO AGENTE DE IA:
    - Todos os checks devem ser verdes (✓)
    - Nenhum X vermelho é permitido
    - Todos os campos devem estar preenchidos
    - Observações gerais devem ser coerentes e substantivas
    """
    result = {
        'document_type': 'Checklist de Credenciamento',
        'is_valid': False,
        'issues': [],
        'warnings': [],
        'details': {},
        'score': 0
    }
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=False)
        sheet = wb.active
        
        # Verificar nome e CNPJ da instituição no cabeçalho
        header_text = ""
        for row in sheet.iter_rows(min_row=1, max_row=10, values_only=True):
            header_text += " ".join([str(cell) for cell in row if cell]) + " "
        
        if institution_name and institution_name.lower() not in header_text.lower():
            result['warnings'].append(f"Nome da instituição '{institution_name}' não encontrado no cabeçalho")
        else:
            result['score'] += 10
            result['details']['institution_found'] = True
        
        if institution_cnpj and institution_cnpj not in header_text:
            result['warnings'].append(f"CNPJ '{institution_cnpj}' não encontrado no cabeçalho")
        else:
            result['score'] += 10
            result['details']['cnpj_found'] = True
        
        # NOVA ANÁLISE: Verificar checkmarks com análise de cores
        green_checks = 0
        red_x_marks = 0
        empty_fields = 0
        filled_fields = 0
        
        for row in sheet.iter_rows(min_row=5):
            for cell in row:
                cell_color = get_cell_background_color(cell)
                cell_value = cell.value
                
                # Verificar se é um campo que deveria estar preenchido
                if cell_value:
                    filled_fields += 1
                    
                    # Verificar símbolos de check ou X
                    if is_check_symbol(cell_value):
                        # Verificar se é verde (aprovado)
                        if cell_color and check_excel_cell_color(cell_color, 'green_check'):
                            green_checks += 1
                        else:
                            green_checks += 1  # Check sem cor ainda conta
                    
                    elif is_x_symbol(cell_value):
                        # X é sempre reprovado
                        red_x_marks += 1
                
                # Verificar células com cor de fundo significativa mas vazias
                elif cell_color:
                    if check_excel_cell_color(cell_color, 'cadprev_answer_yellow') or \
                       check_excel_cell_color(cell_color, 'termo_institution_orange'):
                        empty_fields += 1
        
        result['details']['green_checks'] = green_checks
        result['details']['red_x_marks'] = red_x_marks
        result['details']['filled_fields'] = filled_fields
        result['details']['empty_fields'] = empty_fields
        
        # REGRA CRÍTICA: Nenhum X vermelho permitido
        if red_x_marks > 0:
            result['issues'].append(f"❌ CRÍTICO: Encontrados {red_x_marks} itens marcados com 'X' (REPROVADOS). Todos os itens devem ser ✓")
            result['score'] = 0
            result['is_valid'] = False
            return result
        else:
            result['score'] += 30
        
        if green_checks < 5:
            result['warnings'].append(f"Poucos itens marcados com ✓ encontrados ({green_checks}). Verifique se o checklist está completo.")
        else:
            result['score'] += 20
        
        if empty_fields > 3:
            result['issues'].append(f"❌ Encontrados {empty_fields} campos obrigatórios vazios")
        
        # Procurar por campo "Observações Gerais" ou "Diferenciais Competitivos"
        observacoes_text = ""
        for row in sheet.iter_rows(values_only=True):
            row_text = " ".join([str(cell) for cell in row if cell])
            if "observa" in row_text.lower() or "diferencial" in row_text.lower():
                observacoes_text = row_text
                break
        
        if observacoes_text:
            # Verificar qualidade das observações
            if len(observacoes_text) < 30:
                result['issues'].append("❌ Observações gerais/diferenciais competitivos estão muito rasas (menos de 30 caracteres)")
            elif any(word in observacoes_text.lower() for word in ['teste', 'exemplo', 'aaa', 'xxx', 'asdf']):
                result['issues'].append("❌ Observações gerais parecem ser texto de placeholder/teste - não são coerentes")
            else:
                result['score'] += 30
                result['details']['observations_quality'] = 'adequada'
        else:
            result['warnings'].append("⚠️ Campo de observações gerais/diferenciais competitivos não encontrado")
            result['details']['observations_quality'] = 'não encontrado'
        
        # Verificar estrutura do documento
        total_rows = sheet.max_row
        total_cols = sheet.max_column
        
        if total_rows < 10 or total_cols < 3:
            result['issues'].append("❌ Estrutura do documento parece incompleta (poucas linhas ou colunas)")
        
        wb.close()
        
        # Decisão final
        result['is_valid'] = len(result['issues']) == 0 and result['score'] >= 50
        
    except Exception as e:
        result['issues'].append(f"❌ Erro ao analisar planilha: {str(e)}")
    
    return result

# ============================================
# 3. ANÁLISE: INFORMAÇÕES CADPREV (EXCEL)
# ============================================
def analyze_cadprev(file_path):
    """
    Analisa informações de preenchimento Cadprev (Excel) com CONHECIMENTO ESPECIALIZADO
    
    REGRAS DO AGENTE DE IA:
    - Células AZUIS = Perguntas/Campos (não precisam ser preenchidas)
    - Células AMARELAS = Respostas (DEVEM estar preenchidas)
    - Todas as respostas devem ser coerentes
    """
    result = {
        'document_type': 'Informações Cadprev',
        'is_valid': False,
        'issues': [],
        'warnings': [],
        'details': {},
        'score': 0
    }
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=False)
        sheet = wb.active
        
        # Contadores de cores
        blue_questions = 0
        yellow_answers_filled = 0
        yellow_answers_empty = 0
        all_text = ""
        answers_content = []
        
        for row in sheet.iter_rows():
            for cell in row:
                cell_color = get_cell_background_color(cell)
                cell_value = str(cell.value).strip() if cell.value else ""
                all_text += cell_value + " "
                
                if cell_color:
                    # Verificar células azuis (perguntas)
                    if check_excel_cell_color(cell_color, 'cadprev_question_blue'):
                        blue_questions += 1
                    
                    # Verificar células amarelas (respostas - DEVEM ser preenchidas)
                    elif check_excel_cell_color(cell_color, 'cadprev_answer_yellow'):
                        if cell_value:
                            yellow_answers_filled += 1
                            answers_content.append(cell_value)
                        else:
                            yellow_answers_empty += 1
        
        result['details']['blue_questions'] = blue_questions
        result['details']['yellow_answers_filled'] = yellow_answers_filled
        result['details']['yellow_answers_empty'] = yellow_answers_empty
        
        # REGRA CRÍTICA: Todas as células amarelas devem estar preenchidas
        if yellow_answers_empty > 0:
            result['issues'].append(f"❌ CRÍTICO: {yellow_answers_empty} campos de resposta (amarelos) estão VAZIOS. Todas as respostas devem ser preenchidas.")
            result['score'] = max(0, result['score'] - 30)
        else:
            result['score'] += 40
        
        # Verificar qualidade das respostas
        short_answers = sum(1 for ans in answers_content if len(ans) < 5)
        if short_answers > 3:
            result['warnings'].append(f"⚠️ {short_answers} respostas muito curtas encontradas. Verifique completude.")
        else:
            result['score'] += 20
        
        # Verificar coerência (não placeholders)
        placeholder_patterns = ['teste', 'exemplo', 'xxx', 'aaa', 'preencher', 'inserir']
        incoherent = sum(1 for ans in answers_content if any(p in ans.lower() for p in placeholder_patterns))
        if incoherent > 0:
            result['issues'].append(f"❌ {incoherent} respostas parecem ser placeholders ou texto de teste")
        else:
            result['score'] += 20
        
        # Verificar campos informativos importantes
        important_fields = ['volume', 'gerido', 'cnpj', 'razão', 'instituição']
        found_fields = sum(1 for f in important_fields if f in all_text.lower())
        
        if found_fields < 3:
            result['warnings'].append("⚠️ Alguns campos informativos importantes podem estar faltando")
        else:
            result['score'] += 20
        
        wb.close()
        
        # Decisão final
        result['is_valid'] = len(result['issues']) == 0 and result['score'] >= 50
        
    except Exception as e:
        result['issues'].append(f"❌ Erro ao analisar planilha: {str(e)}")
    
    return result

# ============================================
# 4. ANÁLISE: TERMO DE CREDENCIAMENTO (EXCEL)
# ============================================
def analyze_termo_credenciamento(file_path):
    """
    Analisa Formulário Termo de Credenciamento (Excel)
    - Verifica estrutura
    - Verifica completude informacional
    - Identifica campos que a diretoria deve preencher
    
    REGRAS DO AGENTE DE IA:
    - Células LARANJA/PÊSSEGO = Instituição DEVE preencher (OBRIGATÓRIO)
    - Células BRANCAS = RPPS preenche (apenas reportar se vazios)
    - Campos do RPPS não são obrigatórios para a instituição
    """
    result = {
        'document_type': 'Termo de Credenciamento',
        'is_valid': False,
        'issues': [],
        'warnings': [],
        'details': {},
        'score': 0
    }
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=False)
        sheet = wb.active
        
        # Campos que são responsabilidade do RPPS (NÃO exigir da instituição)
        rpps_fields = [
            'local e data', 'responsável pelo credenciamento', 'número do processo',
            'número do termo', 'ente federativo', 'termo de análise', 'parecer',
            'assinatura do rpps', 'comitê'
        ]
        
        # Contadores
        orange_filled = 0
        orange_empty = 0
        white_empty = 0
        all_text = ""
        institution_responses = []
        
        for row in sheet.iter_rows():
            for cell in row:
                cell_color = get_cell_background_color(cell)
                cell_value = str(cell.value).strip() if cell.value else ""
                all_text += cell_value + " "
                
                if cell_color:
                    # Verificar células laranja (instituição DEVE preencher)
                    if check_excel_cell_color(cell_color, 'termo_institution_orange'):
                        if cell_value:
                            orange_filled += 1
                            institution_responses.append(cell_value)
                        else:
                            orange_empty += 1
                    
                    # Células brancas/claras (RPPS preenche)
                    elif check_excel_cell_color(cell_color, 'termo_rpps_white'):
                        # Verificar se é campo do RPPS (não contar como problema)
                        is_rpps_field = any(f in all_text.lower()[-200:] for f in rpps_fields)
                        if not cell_value and not is_rpps_field:
                            white_empty += 1
        
        result['details']['orange_filled'] = orange_filled
        result['details']['orange_empty'] = orange_empty
        result['details']['white_empty'] = white_empty
        
        # REGRA CRÍTICA: Todos os campos laranja devem estar preenchidos
        if orange_empty > 0:
            result['issues'].append(f"❌ CRÍTICO: {orange_empty} campos obrigatórios da instituição (laranja) estão VAZIOS")
            result['score'] = max(0, result['score'] - 40)
        else:
            result['score'] += 40
        
        if orange_filled < 5:
            result['warnings'].append("⚠️ Poucos campos preenchidos pela instituição")
        else:
            result['score'] += 30
        
        # Reportar campos brancos vazios (para RPPS preencher)
        if white_empty > 0:
            result['details']['rpps_pending_fields'] = white_empty
            result['warnings'].append(f"ℹ️ {white_empty} campos em branco para o RPPS preencher posteriormente")
        
        # Verificar profundidade das respostas
        if institution_responses:
            short_responses = sum(1 for r in institution_responses if len(r) < 10)
            if short_responses > len(institution_responses) * 0.5:
                result['warnings'].append("⚠️ Muitas respostas curtas/superficiais. Verifique completude.")
            else:
                result['score'] += 30
        
        wb.close()
        
        result['is_valid'] = len(result['issues']) == 0 and result['score'] >= 60
        
    except Exception as e:
        result['issues'].append(f"❌ Erro ao analisar documento: {str(e)}")
    
    return result

# ============================================
# 5. ANÁLISE: DECLARAÇÃO UNIFICADA (PDF)
# ============================================
def analyze_declaracao_unificada(file_path):
    """
    Analisa Declaração Unificada (PDF) com CONHECIMENTO ESPECIALIZADO
    
    ⚠️ REGRA CRÍTICA DO AGENTE DE IA:
    - Este documento REQUER assinatura digital VÁLIDA
    - Assinatura inválida = INVALIDA TODO O CREDENCIAMENTO
    - Deve conter texto "DECLARAÇÃO UNIFICADA"
    - Data não pode ser superior a 1 ano
    """
    result = {
        'document_type': 'Declaração Unificada',
        'is_valid': False,
        'issues': [],
        'warnings': [],
        'details': {},
        'score': 0,
        'signature_critical': True  # Marca este documento como crítico para assinatura
    }
    
    try:
        text = extract_text_from_pdf(file_path)
        
        if not text or len(text.strip()) < 50:
            result['issues'].append("❌ Documento vazio ou com muito pouco conteúdo")
            return result
        
        # Verificar se é declaração unificada
        if 'declaração unificada' in text.lower() or 'declaracao unificada' in text.lower():
            result['details']['is_declaracao_unificada'] = True
            result['score'] += 40
        else:
            result['issues'].append("❌ Documento não contém o texto 'DECLARAÇÃO UNIFICADA' - não é o documento correto")
            return result
        
        # Extrair e verificar data
        dates = extract_dates_from_text(text)
        
        if dates:
            most_recent_date = max(dates)
            result['details']['date_found'] = most_recent_date.strftime('%d/%m/%Y')
            
            if is_date_within_one_year(most_recent_date):
                result['details']['date_valid'] = True
                result['score'] += 30
            else:
                result['issues'].append(f"❌ Data da declaração ({most_recent_date.strftime('%d/%m/%Y')}) tem mais de 1 ano - INVÁLIDA")
                result['details']['date_valid'] = False
        else:
            result['warnings'].append("⚠️ Não foi possível identificar a data da declaração")
            result['details']['date_found'] = None
        
        # VALIDAR ASSINATURA DIGITAL (CRÍTICO)
        signature_result = validate_signature_if_applicable(file_path, 'declaracao_unificada')
        
        if signature_result:
            result['details']['signature_validation'] = signature_result
            
            if signature_result.get('validated'):
                if signature_result.get('is_signed') and signature_result.get('signature_valid'):
                    result['details']['signature_status'] = 'VÁLIDA'
                    result['score'] += 30
                    result['warnings'].append("✓ Assinatura digital validada no TCEES")
                elif signature_result.get('is_signed') and not signature_result.get('signature_valid'):
                    # CRÍTICO: Assinatura inválida invalida o credenciamento
                    result['issues'].append("❌ CRÍTICO: Assinatura digital INVÁLIDA - INVALIDA O CREDENCIAMENTO")
                    result['details']['signature_status'] = 'INVÁLIDA - INVALIDA CREDENCIAMENTO'
                    result['score'] = 0
                    result['is_valid'] = False
                    return result
                else:
                    result['issues'].append("❌ CRÍTICO: Documento NÃO possui assinatura digital - OBRIGATÓRIA para Declaração Unificada")
                    result['details']['signature_status'] = 'AUSENTE'
                    result['score'] = max(0, result['score'] - 50)
            else:
                result['warnings'].append(f"⚠️ Validação de assinatura: {signature_result.get('reason', 'indisponível')}")
                result['details']['signature_status'] = 'NÃO VALIDADA'
        
        result['is_valid'] = len(result['issues']) == 0 and result['score'] >= 50
        
    except Exception as e:
        result['issues'].append(f"❌ Erro ao analisar documento: {str(e)}")
    
    return result

# ============================================
# 6. ANÁLISE: RELATÓRIO DE RATING (PDF)
# ============================================
def analyze_relatorio_rating(file_path, institution_name):
    """
    Analisa Relatório de Agência de Risco (Rating) com CONHECIMENTO ESPECIALIZADO
    
    REGRAS DO AGENTE DE IA:
    - Deve conter termos relacionados a rating/risco
    - Deve identificar a agência que emitiu o relatório
    - Deve ter uma classificação/score/nota
    - Deve mencionar a instituição sendo avaliada
    """
    # Importar conhecimento sobre agências de rating
    from ai_document_knowledge import RELATORIO_RATING
    
    result = {
        'document_type': 'Relatório de Rating',
        'is_valid': False,
        'issues': [],
        'warnings': [],
        'details': {},
        'score': 0
    }
    
    try:
        text = extract_text_from_pdf(file_path)
        
        if not text or len(text.strip()) < 100:
            result['issues'].append("❌ Documento vazio ou com muito pouco conteúdo")
            return result
        
        text_lower = text.lower()
        
        # Usar indicadores da base de conhecimento
        rating_indicators = RELATORIO_RATING.get('rating_indicators', [])
        known_agencies = RELATORIO_RATING.get('known_agencies', [])
        rating_patterns = RELATORIO_RATING.get('rating_patterns', [])
        
        # Verificar termos relacionados a rating/risco
        found_terms = [term for term in rating_indicators if term.lower() in text_lower]
        
        if len(found_terms) >= 2:
            result['details']['rating_terms_found'] = found_terms
            result['score'] += 25
        else:
            result['issues'].append("❌ Documento não parece ser um relatório de rating/risco (termos não encontrados)")
            return result
        
        # Identificar a agência que emitiu o relatório
        found_agency = None
        for agency in known_agencies:
            if agency.lower() in text_lower:
                found_agency = agency
                break
        
        if found_agency:
            result['details']['agency_name'] = found_agency
            result['score'] += 25
        else:
            result['warnings'].append("⚠️ Não foi possível identificar a agência de rating")
        
        # Verificar se menciona a instituição avaliada
        if institution_name:
            name_parts = [p for p in institution_name.split() if len(p) > 3]
            if any(part.lower() in text_lower for part in name_parts):
                result['details']['institution_mentioned'] = True
                result['score'] += 20
            else:
                result['warnings'].append(f"⚠️ Nome da instituição '{institution_name}' não encontrado no relatório")
        
        # Verificar se contém classificação/score/nota
        found_rating = None
        for pattern in rating_patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                found_rating = match.group(0)
                break
        
        if found_rating:
            result['details']['rating_found'] = found_rating
            result['details']['has_classification'] = True
            result['score'] += 30
        else:
            result['warnings'].append("⚠️ Não foi possível identificar claramente uma classificação/nota/score")
            result['score'] += 10  # Pontuação parcial
        
        result['is_valid'] = result['score'] >= 50
        
    except Exception as e:
        result['issues'].append(f"❌ Erro ao analisar documento: {str(e)}")
    
    return result

# ============================================
# 7. ANÁLISE: CERTIDÕES
# ============================================
def analyze_certidao(file_path, expected_type):
    """
    Analisa certidões diversas com CONHECIMENTO ESPECIALIZADO
    
    REGRAS DO AGENTE DE IA - Tipos de Certidões:
    - Formulário de Referência CVM: registro na CVM
    - Certidão Autorização Funcionar BACEN: autorização do BC
    - Certidão BACEN Sócio/Representante: dados de sócios no BC
    - Deve ser coerente com o tipo esperado
    - Deve ter formato de documento oficial
    """
    # Importar conhecimento sobre certidões
    from ai_document_knowledge import CERTIDOES
    
    result = {
        'document_type': f'Certidão - {expected_type}',
        'is_valid': False,
        'issues': [],
        'warnings': [],
        'details': {},
        'score': 0
    }
    
    try:
        text = extract_text_from_pdf(file_path)
        
        if not text or len(text.strip()) < 50:
            result['issues'].append("❌ Documento vazio ou com muito pouco conteúdo")
            return result
        
        text_lower = text.lower()
        expected_lower = expected_type.lower()
        
        # Usar conhecimento da base sobre tipos de certidões
        certidao_types = CERTIDOES.get('types', {})
        certificate_indicators = CERTIDOES.get('certificate_indicators', [])
        
        # Tentar identificar o tipo específico de certidão
        matched_type = None
        keywords_to_check = []
        
        for type_key, type_info in certidao_types.items():
            if type_key in expected_lower or any(kw in expected_lower for kw in type_info.get('keywords', [])[:2]):
                matched_type = type_info
                keywords_to_check = type_info.get('keywords', [])
                break
        
        # Fallback para termos genéricos
        if not keywords_to_check:
            keywords_to_check = [word for word in expected_lower.split() if len(word) > 3]
        
        # Verificar palavras-chave
        found_keywords = [kw for kw in keywords_to_check if kw.lower() in text_lower]
        
        if len(found_keywords) >= 2:
            result['details']['keywords_found'] = found_keywords
            result['details']['matches_expected_type'] = True
            result['score'] += 50
        elif len(found_keywords) == 1:
            result['details']['keywords_found'] = found_keywords
            result['warnings'].append(f"⚠️ Apenas 1 palavra-chave encontrada. Verifique se é o tipo correto.")
            result['score'] += 30
        else:
            result['issues'].append(f"❌ Certidão não parece ser do tipo '{expected_type}'. Palavras-chave não encontradas.")
        
        # Verificar se tem formato de certidão oficial
        cert_format_found = sum(1 for ind in certificate_indicators if ind.lower() in text_lower)
        
        if cert_format_found >= 2:
            result['details']['looks_like_certificate'] = True
            result['score'] += 40
        elif cert_format_found == 1:
            result['score'] += 20
            result['warnings'].append("⚠️ Documento pode não ter formato de certidão oficial")
        else:
            result['warnings'].append("⚠️ Documento não parece ter formato de certidão oficial")
        
        # Adicionar descrição do tipo se encontrado
        if matched_type:
            result['details']['expected_description'] = matched_type.get('description', '')
        
        result['is_valid'] = result['score'] >= 50
        
    except Exception as e:
        result['issues'].append(f"❌ Erro ao analisar documento: {str(e)}")
    
    return result

# ============================================
# FUNÇÃO PRINCIPAL DE ANÁLISE
# ============================================
def analyze_document(file_path, document_type, document_name, institution_name=None, institution_cnpj=None):
    """
    Função principal que roteia para o analisador apropriado
    Usa a base de conhecimento do agente de IA para análise inteligente
    
    REGRA GLOBAL DE ASSINATURA:
    - Se documento PDF tem assinatura, validar no TCEES
    - Declaração Unificada: assinatura inválida = INVALIDA credenciamento
    - Outros documentos: apenas reportar no feedback
    """
    
    print(f"\n🔍 analyze_document() chamado:")
    print(f"   - Tipo: {document_type}")
    print(f"   - Nome: {document_name}")
    print(f"   - Instituição: {institution_name}")
    
    # Normalizar tipo de documento
    doc_type_lower = document_type.lower()
    
    # Roteamento baseado no tipo
    result = None
    
    if 'apresentacao' in doc_type_lower or 'institucional' in doc_type_lower:
        print(f"   ➡️  Roteando para: analyze_apresentacao_institucional")
        result = analyze_apresentacao_institucional(file_path, institution_name)
    
    elif 'checklist' in doc_type_lower:
        print(f"   ➡️  Roteando para: analyze_checklist_credenciamento")
        result = analyze_checklist_credenciamento(file_path, institution_name, institution_cnpj)
    
    elif 'cadprev' in doc_type_lower:
        print(f"   ➡️  Roteando para: analyze_cadprev")
        result = analyze_cadprev(file_path)
    
    elif 'termo' in doc_type_lower and 'credenciamento' in doc_type_lower:
        print(f"   ➡️  Roteando para: analyze_termo_credenciamento")
        result = analyze_termo_credenciamento(file_path)
    
    elif 'declaracao' in doc_type_lower and 'unificada' in doc_type_lower:
        print(f"   ➡️  Roteando para: analyze_declaracao_unificada")
        result = analyze_declaracao_unificada(file_path)
    
    elif 'rating' in doc_type_lower or 'risco' in doc_type_lower:
        print(f"   ➡️  Roteando para: analyze_relatorio_rating")
        result = analyze_relatorio_rating(file_path, institution_name)
    
    elif 'certidao' in doc_type_lower or 'certidão' in doc_type_lower or 'cvm' in doc_type_lower or 'bacen' in doc_type_lower:
        print(f"   ➡️  Roteando para: analyze_certidao")
        result = analyze_certidao(file_path, document_name)
    
    else:
        print(f"   ⚠️  Tipo não reconhecido - análise genérica com validação de assinatura")
        result = {
            'document_type': document_name,
            'is_valid': True,
            'issues': [],
            'warnings': ['⚠️ Tipo de documento não reconhecido. Análise especializada não aplicada.'],
            'details': {},
            'score': 50
        }
    
    # REGRA GLOBAL: Validar assinatura para documentos PDF que não foram validados internamente
    if result and file_path.lower().endswith('.pdf'):
        # Se não tem informação de assinatura, validar
        if 'signature_validation' not in result.get('details', {}):
            signature_result = validate_signature_if_applicable(file_path, document_type)
            
            if signature_result and signature_result.get('validated'):
                result['details'] = result.get('details', {})
                result['details']['signature_validation'] = signature_result
                
                if signature_result.get('is_signed'):
                    if signature_result.get('signature_valid'):
                        result['warnings'] = result.get('warnings', [])
                        result['warnings'].append("✓ Assinatura digital validada no TCEES")
                        result['details']['signature_status'] = 'VÁLIDA'
                    else:
                        # Verificar se é documento crítico
                        if is_signature_critical(document_type):
                            result['issues'] = result.get('issues', [])
                            result['issues'].append("❌ CRÍTICO: Assinatura digital INVÁLIDA - INVALIDA O CREDENCIAMENTO")
                            result['is_valid'] = False
                            result['score'] = 0
                            result['details']['signature_status'] = 'INVÁLIDA - CRÍTICO'
                        else:
                            result['warnings'] = result.get('warnings', [])
                            result['warnings'].append("⚠️ Assinatura digital inválida detectada")
                            result['details']['signature_status'] = 'INVÁLIDA'
    
    return result


# ============================================
# FUNÇÃO DE ANÁLISE COMPLETA COM IA
# ============================================
def analyze_document_with_ai(file_path, document_type, document_name, institution_name=None, institution_cnpj=None):
    """
    Análise completa do documento usando IA + regras do agente
    Combina análise de regras com análise semântica de IA
    """
    
    # Primeiro, fazer análise baseada em regras
    rules_result = analyze_document(file_path, document_type, document_name, institution_name, institution_cnpj)
    
    # Se IA disponível, enriquecer com análise semântica
    ai_status = get_ai_status()
    
    if ai_status.get('available'):
        try:
            # Extrair conteúdo do documento
            if file_path.lower().endswith('.pdf'):
                content = extract_text_from_pdf(file_path)
            elif file_path.lower().endswith(('.xlsx', '.xlsm', '.xls')):
                content = extract_excel_content_for_ai(file_path)
            else:
                content = ""
            
            if content and len(content) > 100:
                # Obter prompt especializado da base de conhecimento
                ai_prompt = get_ai_prompt_for_document(
                    document_type,
                    institution_name=institution_name or '',
                    institution_cnpj=institution_cnpj or '',
                    expected_type=document_name
                )
                
                if ai_prompt:
                    ai_result = get_ai_analysis(ai_prompt, content, document_type)
                    
                    if ai_result.get('success'):
                        rules_result['ai_content_analysis'] = ai_result.get('analysis', {})
                        rules_result['details'] = rules_result.get('details', {})
                        rules_result['details']['ai_powered'] = True
                        rules_result['details']['ai_provider'] = ai_result.get('provider', 'Desconhecido')
                        
                        # Combinar score da IA com score das regras
                        ai_score = rules_result['ai_content_analysis'].get('score', 0)
                        if ai_score:
                            rules_result['score'] = int((rules_result['score'] + ai_score) / 2)
                            
                            # Se IA invalidou, considerar
                            if not rules_result['ai_content_analysis'].get('is_valid', True):
                                ai_issues = rules_result['ai_content_analysis'].get('issues', [])
                                rules_result['issues'] = rules_result.get('issues', []) + ai_issues
                                if ai_issues:
                                    rules_result['is_valid'] = False
        
        except Exception as e:
            print(f"⚠️ Erro na análise de IA: {e}")
            rules_result['details'] = rules_result.get('details', {})
            rules_result['details']['ai_error'] = str(e)
    
    return rules_result
